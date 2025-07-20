import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
from appdirs import user_data_dir
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import mplfinance as mpf

class BreakoutExporter:
    def __init__(self, output_dir: str = None, filename_base: str = "breakout_candidates"):
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.output_dir = Path(output_dir) if output_dir else Path(user_data_dir('ScreenerEngine', 'Miltonstreet')) / 'exports'
            #self.filename_stem = Path(filename_base).with_name(f"{Path(filename_base).name}_{timestamp}")
            self.filename_stem = Path(filename_base).with_name(f"{timestamp}_{Path(filename_base).name}")
            self.full_base_path = self.output_dir / self.filename_stem
            self.full_base_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            print("❌ Error initializing BreakoutExporter:")
            traceback.print_exc()
            raise

    def _build_output_path(self, extension: str) -> Path:
        return self.full_base_path.with_suffix(f".{extension}")

    def export_to_csv(self, df: pd.DataFrame) -> Path | None:
        try:
            path = self._build_output_path("csv")
            df.to_csv(path, index=False)
            print(f"✅ Exported to CSV: {path}")
            return path
        except Exception:
            print("❌ Failed to export CSV:")
            traceback.print_exc()
            return None

    def export_to_excel(self, df: pd.DataFrame) -> Path | None:
        try:
            path = self._build_output_path("xlsx")
            df.to_excel(path, index=False)

            wb = load_workbook(path)
            ws = wb.active
            ws.freeze_panes = ws["A2"]

            from openpyxl.styles import numbers

            column_format_map = {
                # Existing breakout metrics
                "last_close": numbers.FORMAT_CURRENCY_USD_SIMPLE,
                "sma_20": numbers.FORMAT_CURRENCY_USD_SIMPLE,
                "sma_50": numbers.FORMAT_CURRENCY_USD_SIMPLE,
                "resistance_intercept": numbers.FORMAT_CURRENCY_USD_SIMPLE,
                "support_intercept": numbers.FORMAT_CURRENCY_USD_SIMPLE,

                "market_cap": "#,##0",
                "avg_volume_50": "#,##0",
                "last_volume": "#,##0",

                "pivot_high_count": "0",
                "pivot_low_count": "0",
                "pivot_high_strength_avg": "0.00",
                "pivot_low_strength_avg": "0.00",

                "resistance_r2": "0.00%",
                "support_r2": "0.00%",
                "resistance_slope": "0.00",
                "support_slope": "0.00",

                "bullish_score": "0.00",
                "volume_ratio": "0.00%",

                # New fields from v_sound_breakout_candidates
                "pct_from_20d_high": "0.00%",
                "pct_range_5d": "0.00%",
                "avg_move_pct": "0.00%",

                # New fields from BreakoutScore
                "score": "0.00",
                "tightness_score": "0.0000",
                "support_slope": "0.0000",
                "stddev_close": "0.0000",

                # Closed trades performance metrics
                "total_bought": "#,##0",
                "total_cost": numbers.FORMAT_CURRENCY_USD_SIMPLE,
                "total_proceeds": numbers.FORMAT_CURRENCY_USD_SIMPLE,
                "net_gain": numbers.FORMAT_CURRENCY_USD_SIMPLE,
                "pct_gain": "0.00%",
                "first_buy_date": "yyyy-mm-dd",
                "last_sell_date": "yyyy-mm-dd",
                "holding_days": "0",
            }

            for col_idx, column_cells in enumerate(ws.columns, 1):
                col_letter = get_column_letter(col_idx)
                header = column_cells[0].value
                max_length = 0
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                        if header in column_format_map and isinstance(cell.value, (int, float)):
                            cell.number_format = column_format_map[header]
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(path)
            print(f"✅ Exported to Excel: {path}")
            return path
        except Exception as e:
            print(f"❌ Failed to export Excel: {e}")
            traceback.print_exc()
            return None

    def export_charts(self, summary_df: pd.DataFrame, history_df: pd.DataFrame, excel_path: Path):
        try:
            output_dir = excel_path.with_suffix('')
            output_dir.mkdir(parents=True, exist_ok=True)

            history_df.columns = history_df.columns.str.lower()

            for i, row in summary_df.iterrows():
                symbol = row['symbol']
                start_date = pd.to_datetime(row['start_date'])
                end_date = pd.to_datetime(row['end_date'])
                r_slope = row['resistance_slope']
                r_intercept = row['resistance_intercept']
                s_slope = row['support_slope']
                s_intercept = row['support_intercept']

                history = history_df[history_df['symbol'] == symbol].copy()
                if history.empty:
                    print(f"⚠️ Skipping {symbol}: no historical data.")
                    continue

                history['date'] = pd.to_datetime(history['date'])
                history.set_index('date', inplace=True)
                history.sort_index(inplace=True)

                if history.empty or len(history) < 2:
                    print(f"⚠️ Skipping {symbol}: not enough data to plot.")
                    continue

                ordinal_dates = (history.index.astype(int) // 10**9).astype(int)
                history['resistance_line'] = r_slope * ordinal_dates + r_intercept
                history['support_line'] = s_slope * ordinal_dates + s_intercept
                history['sma50'] = history['close'].rolling(window=50).mean()

                chart_data = history[['open', 'high', 'low', 'close', 'volume']]

                # Build additional plots only if they have non-NaN values
                add_plots = []
                if history['sma50'].notna().any():
                    add_plots.append(mpf.make_addplot(history['sma50'], color='blue', width=0.8))
                if history['resistance_line'].notna().any():
                    add_plots.append(mpf.make_addplot(history['resistance_line'], color='red', linestyle='--'))
                if history['support_line'].notna().any():
                    add_plots.append(mpf.make_addplot(history['support_line'], color='green', linestyle='--'))

                output_file = output_dir / f"{str(i + 1).zfill(4)}_{symbol}.png"

                mpf.plot(
                    chart_data,
                    type='candle',
                    volume=True,
                    style='yahoo',
                    title=f"{symbol} Breakout Chart",
                    addplot=add_plots,
                    savefig=dict(fname=output_file, dpi=120, bbox_inches='tight')
                )

                print(f"📈 Chart saved: {output_file}")

        except Exception as e:
            print("❌ Failed to export charts:")
            print(f"👉 {e}")
            traceback.print_exc()
